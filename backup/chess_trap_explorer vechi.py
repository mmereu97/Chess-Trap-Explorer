import pygame
import chess
import pandas as pd
import os
import pyperclip
import pygame_textinput
import json

# --- CONSTANTE ȘI CONFIGURARE ---
MARGIN = 30
BOARD_SIZE = 512
SQUARE_SIZE = BOARD_SIZE // 8
BUTTONS_WIDTH = 200  # Nouă zonă pentru butoane în stânga
SIDEBAR_WIDTH = 400  # Mărită zona de capcane
WIDTH = BUTTONS_WIDTH + BOARD_SIZE + 2 * MARGIN + SIDEBAR_WIDTH
HEIGHT = BOARD_SIZE + 2 * MARGIN + 60 # Spațiu extra pentru panoul de istoric/înregistrare
BROWN = (181, 136, 99)
LIGHT_BROWN = (240, 217, 181)
LIVE_DARK = (118, 150, 86)  # Verde închis
LIVE_LIGHT = (238, 238, 210)  # Crem deschis
BACKGROUND_COLOR = (49, 46, 43)
WHITE_PIECE_COLOR = (248, 248, 248)
BLACK_PIECE_COLOR = (88, 88, 88)
OUTLINE_COLOR = (0, 0, 0)
ASSISTANT_TEXT_COLOR = (255, 255, 0)
HIGHLIGHT_COLOR = (255, 0, 0, 100)
COORDINATE_COLOR = (220, 220, 220)
UNDO_BUTTON_COLOR = (100, 100, 200)
BUTTON_COLOR = (80, 80, 150)
BUTTON_TEXT_COLOR = (255, 255, 255)
SELECTED_BUTTON_COLOR = (120, 120, 200)
TARGET_SQUARE_COLOR = (0, 255, 255, 100)
HISTORY_BACKGROUND = (60, 60, 60)
HISTORY_TEXT_COLOR = (200, 200, 200)
RECORD_BUTTON_COLOR = (200, 50, 50)
RECORDING_BUTTON_COLOR = (50, 200, 50)
RESTART_BUTTON_COLOR = (150, 75, 0)
EXPECTED_MOVE_COLOR = (0, 255, 0, 100)  # Verde transparent pentru mutarea așteptată
MESSAGE_BACKGROUND = (40, 40, 50)
MESSAGE_TEXT_COLOR = (220, 220, 220)

pygame.init()
pygame.font.init()
TITLE_FONT = pygame.font.SysFont('Arial', 48, bold=True)
ASSISTANT_FONT = pygame.font.SysFont('Arial', 24, bold=True)
COORDINATE_FONT = pygame.font.SysFont('Arial', 16)
HISTORY_FONT = pygame.font.SysFont('Consolas', 14)
WIN = pygame.display.set_mode((WIDTH, HEIGHT))
pygame.display.set_caption("Antrenor și Înregistrator de Capcane")


def load_settings():
    """Încarcă setările din fișierul settings.json"""
    default_settings = {
        "pgn_import_max_moves": 25,
        "last_pgn_directory": "",
        "window_width": WIDTH,
        "window_height": HEIGHT
    }
    
    try:
        with open("settings.json", "r", encoding="utf-8") as f:
            settings = json.load(f)
            # Verifică că toate setările necesare există
            for key, default_value in default_settings.items():
                if key not in settings:
                    settings[key] = default_value
            return settings
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"[SETĂRI] Nu pot încărca setările, folosesc valorile implicite: {e}")
        save_settings(default_settings)
        return default_settings

def save_settings(settings):
    """Salvează setările în fișierul settings.json"""
    try:
        with open("settings.json", "w", encoding="utf-8") as f:
            json.dump(settings, f, indent=2, ensure_ascii=False)
        print(f"[SETĂRI] Setările au fost salvate.")
    except Exception as e:
        print(f"[EROARE SETĂRI] Nu pot salva setările: {e}")

def update_setting(key, value):
    """Actualizează o singură setare"""
    settings = load_settings()
    settings[key] = value
    save_settings(settings)


# --- FUNCȚII UTILITARE ---
def build_trap_library_from_excel(player_color, file_path="traps.xlsx"):
    trap_library = {}
    total_traps = 0  # ADĂUGAT: Contor pentru capcane
    
    if not os.path.exists(file_path):
        print(f"[EROARE] Fișierul '{file_path}' nu a fost găsit.")
        return {}, 0  # MODIFICAT: Returnează și numărul
    
    sheet_to_read = "White" if player_color == chess.WHITE else "Black"
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_to_read, header=None)
        print(f"[INFO] Se procesează sheet-ul '{sheet_to_read}'...")
        for index, row in df.iloc[1:].iterrows():
            trap_name, moves = row[0], row[1:].dropna().tolist()
            if not moves: continue
            
            total_traps += 1  # ADĂUGAT: Incrementează contorul
            
            board = chess.Board()
            start_fen = board.fen()
            if start_fen not in trap_library: trap_library[start_fen] = {"player_options": []}
            if not any(opt['suggestion'] == moves[0] and opt['name'] == trap_name for opt in trap_library[start_fen]["player_options"]):
                trap_library[start_fen]["player_options"].append({"name": trap_name, "suggestion": moves[0], "opponent_target_san": moves[1] if len(moves) > 1 else None})
            for i, move_san in enumerate(moves):
                try: board.push_san(move_san)
                except ValueError: print(f"[EROARE] Mutare invalidă '{move_san}' în '{trap_name}'."); break
                if board.turn == player_color and i + 1 < len(moves):
                    current_fen, next_suggestion, next_opponent_target = board.fen(), moves[i+1], moves[i+2] if i + 2 < len(moves) else None
                    if current_fen not in trap_library: trap_library[current_fen] = {"player_options": []}
                    if not any(opt['suggestion'] == next_suggestion for opt in trap_library[current_fen]["player_options"]):
                        trap_library[current_fen]["player_options"].append({"name": trap_name, "suggestion": next_suggestion, "opponent_target_san": next_opponent_target})
        print(f"[INFO] Biblioteca a fost construită: {len(trap_library)} poziții din {total_traps} capcane totale.")
        return trap_library, total_traps  # MODIFICAT: Returnează ambele
    except Exception as e: print(f"[EROARE FATALĂ] Nu am putut citi Excel: {e}"); return {}, 0  # MODIFICAT


def write_new_line_to_excel(sheet_name, row_index, trap_name, moves, file_path="traps.xlsx"):
    try:
        from openpyxl import load_workbook
        book = load_workbook(file_path)
        if sheet_name not in book.sheetnames:
            book.create_sheet(sheet_name)
            sheet = book[sheet_name]
            sheet.cell(row=1, column=1, value="denumirea")
            for i in range(30): sheet.cell(row=1, column=i+2, value="white" if i % 2 == 0 else "black")
        else:
            sheet = book[sheet_name]
        for col_idx in range(1, 31): sheet.cell(row=row_index, column=col_idx, value=None)
        sheet.cell(row=row_index, column=1, value=trap_name)
        for i, move in enumerate(moves): sheet.cell(row=row_index, column=i + 2, value=move)
        book.save(file_path)
        print(f"[SUCCES] Linia '{trap_name}' a fost salvată/suprascrisă în '{sheet_name}' pe rândul {row_index}.")
        return True
    except Exception as e: print(f"[EROARE FATALĂ] Nu am putut scrie în Excel: {e}"); return False


def get_next_free_row(sheet_name, file_path="traps.xlsx"):
    """Detectează automat următorul rând liber din sheet"""
    try:
        from openpyxl import load_workbook
        import os
        
        if not os.path.exists(file_path):
            return 2  # Primul rând după header
        
        book = load_workbook(file_path)
        if sheet_name not in book.sheetnames:
            return 2  # Primul rând după header pentru sheet nou
        
        sheet = book[sheet_name]
        
        # Caută primul rând gol pornind de la rândul 2
        row = 2
        while sheet.cell(row=row, column=1).value is not None:
            row += 1
        
        book.close()
        print(f"[AUTO-DETECT] Următorul rând liber în '{sheet_name}': {row}")
        return row
        
    except Exception as e:
        print(f"[EROARE AUTO-DETECT] Nu pot detecta rândul liber: {e}")
        return 2  # Fallback la rândul 2

def sort_traps_by_common_beginning(trap_suggestions, current_position_moves):
    """
    Sortează capcanele în funcție de cât de multe capcane au secvența de început comună.
    Prima prioritate: capcanele cu cea mai lungă secvență comună cu poziția curentă
    A doua prioritate: capcanele care apar în grupuri mai mari
    """
    if not trap_suggestions:
        return trap_suggestions
    
    # Construiește un dicționar cu toate secvențele de capcane pentru analiza grupurilor
    all_sequences = {}
    
    # Parcurge fișierul Excel pentru a obține toate secvențele
    try:
        import pandas as pd
        import os
        
        if os.path.exists("traps.xlsx"):
            # CORECTARE: Asigură-te că player_choice este definit în contextul global
            try:
                sheet_name = "White" if player_choice == chess.WHITE else "Black"
            except NameError:
                # Fallback dacă player_choice nu e disponibil
                return trap_suggestions
            
            df = pd.read_excel("traps.xlsx", sheet_name=sheet_name, header=None)
            
            for index, row in df.iloc[1:].iterrows():  # Skip header
                trap_name, moves = row[0], row[1:].dropna().tolist()
                if moves and trap_name:  # CORECTARE: Verifică că trap_name nu e None
                    all_sequences[str(trap_name)] = moves  # CORECTARE: Convertește la string
    except Exception as e:
        print(f"[SORTARE] Nu pot citi Excel pentru sortare: {e}")
        return trap_suggestions
    
    def calculate_trap_priority(trap):
        trap_name = str(trap.get('name', ''))  # CORECTARE: Convertește la string și gestionează None
        if not trap_name or trap_name not in all_sequences:
            return (0, 0, 0, trap_name)  # CORECTARE: Returnează întotdeauna 4 valori întregi + string
        
        full_sequence = all_sequences[trap_name]
        
        # 1. Calculează cât de mult se potrivește cu mutările curente
        common_length = 0
        min_length = min(len(current_position_moves), len(full_sequence))
        
        for i in range(min_length):
            if current_position_moves[i] == full_sequence[i]:
                common_length += 1
            else:
                break
        
        # 2. Calculează câte capcane au același început ca această capcană
        group_size = 0
        sequence_to_check = full_sequence[:len(current_position_moves) + 1]  # +1 pentru următoarea mutare
        
        for other_name, other_sequence in all_sequences.items():
            if len(other_sequence) >= len(sequence_to_check):
                matches = True
                for i in range(len(sequence_to_check)):
                    if i >= len(other_sequence) or sequence_to_check[i] != other_sequence[i]:
                        matches = False
                        break
                if matches:
                    group_size += 1
        
        # 3. Calculează câte capcane au aceeași următoare mutare
        next_move_group_size = 0
        next_move = trap.get('suggestion', '')
        
        for other_trap in trap_suggestions:
            if other_trap.get('suggestion', '') == next_move:
                next_move_group_size += 1
        
        # CORECTARE: Returnează întotdeauna exact 4 valori întregi + 1 string
        # Toate valorile sunt negative pentru sortare descrescătoare
        return (-int(common_length), -int(next_move_group_size), -int(group_size), str(trap_name))
    
    try:
        # Sortează capcanele folosind criteriile de prioritate
        sorted_traps = sorted(trap_suggestions, key=calculate_trap_priority)
        print(f"[SORTARE] Sortate {len(sorted_traps)} capcane după secvența comună")
        return sorted_traps
    except Exception as e:
        print(f"[EROARE SORTARE] Nu pot sorta capcanele: {e}")
        # Returnează lista originală dacă sortarea eșuează
        print("[SORTARE] Folosesc ordinea originală din Excel")
        return trap_suggestions

def get_current_position_moves(board):
    """Extrage secvența de mutări care a dus la poziția curentă"""
    moves = []
    temp_board = chess.Board()
    
    for move in board.move_stack:
        san_move = temp_board.san(move)
        moves.append(san_move)
        temp_board.push(move)
    
    return moves

def import_traps_from_pgn(pgn_file_path, max_moves=12):
    """Importă capcane din fișier PGN și le returnează organizate pe culori"""
    import chess.pgn
    
    white_traps = []
    black_traps = []
    
    try:
        with open(pgn_file_path, 'r', encoding='utf-8') as pgn_file:
            game_count = 0
            processed_count = 0
            while True:
                game = chess.pgn.read_game(pgn_file)
                if game is None:
                    break
                
                game_count += 1
                moves = list(game.mainline_moves())
                
                print(f"[DEBUG] Partida {game_count}: {len(moves)} mutări, nume: {game.headers.get('ChapterName', 'N/A')}")
                
                # Pentru Lichess Studies - acceptă orice lungime rezonabilă
                if 4 <= len(moves) <= max_moves:
                    # Convertește mutările în notație SAN
                    board = chess.Board()
                    san_moves = []
                    
                    try:
                        for move in moves:
                            san_moves.append(board.san(move))
                            board.push(move)
                        
                        # Pentru study-uri, folosim numele ca indicator
                        trap_name = game.headers.get("ChapterName", game.headers.get("Event", f"Trapă {game_count}"))
                        
                        # Verifică dacă se termină cu mat
                        is_checkmate = "#" in san_moves[-1] if san_moves else False
                        if is_checkmate:
                            trap_name += " (Mat)"
                        
                        print(f"[DEBUG] Mutări SAN: {san_moves}")
                        print(f"[DEBUG] Este mat: {is_checkmate}")
                        
                        # Determină cine câștigă bazat pe numărul de mutări
                        # Pentru study-uri, presupunem că cel care dă ultimul mov câștigă
                        total_moves = len(san_moves)
                        if total_moves > 0:
                            if is_checkmate:
                                # Dacă e mat, cel care a mutat ultimul câștigă
                                if total_moves % 2 == 1:  # Mutare impară = albele au mutat ultimul
                                    white_traps.append({
                                        "name": trap_name,
                                        "moves": san_moves,
                                        "result": "1-0"
                                    })
                                    processed_count += 1
                                    print(f"[DEBUG] Adăugat în capcane albe: {trap_name}")
                                else:  # Mutare pară = negrele au mutat ultimul
                                    black_traps.append({
                                        "name": trap_name,
                                        "moves": san_moves,
                                        "result": "0-1"
                                    })
                                    processed_count += 1
                                    print(f"[DEBUG] Adăugat în capcane negre: {trap_name}")
                            else:
                                # Dacă nu e mat, dar e o trapă, încercăm să o adăugăm ca trapă pentru albe
                                # (majoritatea trap-urilor din study-uri sunt pentru albe)
                                white_traps.append({
                                    "name": trap_name,
                                    "moves": san_moves,
                                    "result": "1-0"
                                })
                                processed_count += 1
                                print(f"[DEBUG] Adăugat în capcane albe (non-mat): {trap_name}")
                    
                    except Exception as e:
                        print(f"[DEBUG] Eroare la procesarea partidei {game_count}: {e}")
                        continue
                else:
                    if len(moves) > 0:
                        print(f"[DEBUG] Partida ignorată - {len(moves)} mutări (limita: {max_moves})")
        
        print(f"[IMPORT] Am procesat {game_count} partide din PGN.")
        print(f"[IMPORT] Procesate cu succes: {processed_count}")
        print(f"[IMPORT] Găsite: {len(white_traps)} capcane pentru albe, {len(black_traps)} pentru negre.")
        
    except Exception as e:
        print(f"[EROARE IMPORT] Nu pot citi fișierul PGN: {e}")
        
    return white_traps, black_traps

def save_traps_to_excel(white_traps, black_traps, file_path="traps.xlsx"):
    """Salvează capcanele importate în Excel cu verificare de duplicate"""
    try:
        from openpyxl import load_workbook, Workbook
        import os
        
        # Încarcă sau creează workbook
        if os.path.exists(file_path):
            book = load_workbook(file_path)
        else:
            book = Workbook()
            if "Sheet" in book.sheetnames:
                del book["Sheet"]
        
        def check_for_duplicates_and_add(traps, sheet_name):
            if not traps:
                return 0
                
            if sheet_name not in book.sheetnames:
                book.create_sheet(sheet_name)
            sheet = book[sheet_name]
            
            # Citește capcanele existente pentru comparație
            existing_names = set()
            existing_sequences = set()
            
            row = 2  # Începe de la rândul 2 (rândul 1 e header)
            while sheet.cell(row=row, column=1).value is not None:
                name = sheet.cell(row=row, column=1).value
                if name:
                    existing_names.add(name.lower())
                    
                    # Construiește secvența de mutări existentă
                    moves = []
                    col = 2
                    while sheet.cell(row=row, column=col).value is not None:
                        moves.append(sheet.cell(row=row, column=col).value)
                        col += 1
                    if moves:
                        existing_sequences.add(tuple(moves))
                row += 1
            
            # Găsește primul rând liber
            next_row = row
            added_count = 0
            
            for trap in traps:
                moves_tuple = tuple(trap["moves"])
                original_name = trap["name"]
                
                # Verifică dacă secvența de mutări există deja
                if moves_tuple in existing_sequences:
                    print(f"[SKIP] Secvența identică găsită, sărim: {original_name}")
                    continue
                
                # Verifică dacă numele există și creează unul unic
                final_name = original_name
                name_lower = original_name.lower()
                counter = 1
                
                while name_lower in existing_names:
                    final_name = f"{original_name} ({counter})"
                    name_lower = final_name.lower()
                    counter += 1
                
                # Adaugă numele în set-ul de nume existente
                existing_names.add(name_lower)
                existing_sequences.add(moves_tuple)
                
                # Salvează capcana cu numele final
                sheet.cell(row=next_row, column=1, value=final_name)
                for i, move in enumerate(trap["moves"]):
                    sheet.cell(row=next_row, column=i + 2, value=move)
                
                if final_name != original_name:
                    print(f"[RENAME] '{original_name}' -> '{final_name}'")
                else:
                    print(f"[ADD] Adăugat: {final_name}")
                
                next_row += 1
                added_count += 1
            
            return added_count
        
        # Procesează capcanele pentru ambele culori
        white_added = check_for_duplicates_and_add(white_traps, "White")
        black_added = check_for_duplicates_and_add(black_traps, "Black")
        
        book.save(file_path)
        print(f"[SUCCES] Salvate {white_added} capcane albe și {black_added} capcane negre în {file_path}")
        
        if white_added == 0 and black_added == 0:
            print("[INFO] Nu s-au adăugat capcane noi - toate erau duplicate.")
        
        return True
        
    except Exception as e:
        print(f"[EROARE SALVARE] Nu pot salva în Excel: {e}")
        return False

def start_screen():
    selected_color = chess.WHITE
    white_button_rect = pygame.Rect(WIDTH / 2 - 150, HEIGHT / 2 - 50, 300, 50)
    black_button_rect = pygame.Rect(WIDTH / 2 - 150, HEIGHT / 2 + 10, 300, 50)
    start_button_rect = pygame.Rect(WIDTH / 2 - 100, HEIGHT / 2 + 100, 200, 50)
    running = True
    while running:
        WIN.fill(BACKGROUND_COLOR)
        title_text = TITLE_FONT.render("Antrenor de Șah", True, WHITE_PIECE_COLOR); WIN.blit(title_text, (WIDTH / 2 - title_text.get_width() / 2, HEIGHT / 4))
        pygame.draw.rect(WIN, SELECTED_BUTTON_COLOR if selected_color == chess.WHITE else BUTTON_COLOR, white_button_rect, border_radius=8); white_text = ASSISTANT_FONT.render("Joc cu Albele", True, BUTTON_TEXT_COLOR); WIN.blit(white_text, (white_button_rect.x + (white_button_rect.width - white_text.get_width()) / 2, white_button_rect.y + 10))
        pygame.draw.rect(WIN, SELECTED_BUTTON_COLOR if selected_color == chess.BLACK else BUTTON_COLOR, black_button_rect, border_radius=8); black_text = ASSISTANT_FONT.render("Joc cu Negrele", True, BUTTON_TEXT_COLOR); WIN.blit(black_text, (black_button_rect.x + (black_button_rect.width - black_text.get_width()) / 2, black_button_rect.y + 10))
        pygame.draw.rect(WIN, UNDO_BUTTON_COLOR, start_button_rect, border_radius=8); start_text = ASSISTANT_FONT.render("Start Joc", True, BUTTON_TEXT_COLOR); WIN.blit(start_text, (start_button_rect.x + (start_button_rect.width - start_text.get_width()) / 2, start_button_rect.y + 10))
        for event in pygame.event.get():
            if event.type == pygame.QUIT: return None
            if event.type == pygame.MOUSEBUTTONDOWN:
                if white_button_rect.collidepoint(event.pos): selected_color = chess.WHITE; print("[CONFIG] Ai ales să joci cu Albele.")
                elif black_button_rect.collidepoint(event.pos): selected_color = chess.BLACK; print("[CONFIG] Ai ales să joci cu Negrele.")
                elif start_button_rect.collidepoint(event.pos): print(f"[CONFIG] Pornesc jocul..."); return selected_color
        pygame.display.flip()

def load_piece_images():
    """Încarcă imaginile pieselor din folderul pieces/"""
    images = {}
    piece_files = {
        'wK': 'pieces/wK.png', 'wQ': 'pieces/wQ.png', 'wR': 'pieces/wR.png',
        'wB': 'pieces/wB.png', 'wN': 'pieces/wN.png', 'wP': 'pieces/wP.png',
        'bK': 'pieces/bK.png', 'bQ': 'pieces/bQ.png', 'bR': 'pieces/bR.png',
        'bB': 'pieces/bB.png', 'bN': 'pieces/bN.png', 'bP': 'pieces/bP.png'
    }
    
    for piece_key, file_path in piece_files.items():
        try:
            # Încarcă imaginea
            original_image = pygame.image.load(file_path).convert_alpha()
            # Redimensionează la mărimea pătratului
            scaled_image = pygame.transform.scale(original_image, (SQUARE_SIZE, SQUARE_SIZE))
            images[piece_key] = scaled_image
            print(f"[ÎNCĂRCARE] Piesa {piece_key} încărcată cu succes")
        except pygame.error as e:
            print(f"[EROARE] Nu pot încărca {file_path}: {e}")
            # Fallback la piese text simple
            piece_surface = pygame.Surface((SQUARE_SIZE, SQUARE_SIZE), pygame.SRCALPHA)
            color = WHITE_PIECE_COLOR if piece_key.startswith('w') else BLACK_PIECE_COLOR
            text = piece_key[1]  # Doar litera (K, Q, R, etc.)
            font = pygame.font.Font(None, SQUARE_SIZE // 2)
            text_surface = font.render(text, True, color)
            text_rect = text_surface.get_rect(center=(SQUARE_SIZE//2, SQUARE_SIZE//2))
            piece_surface.blit(text_surface, text_rect)
            images[piece_key] = piece_surface
    
    return images

def draw_coordinates(win, flipped):
    files, ranks = "abcdefgh", "87654321"
    if flipped: files = files[::-1]; ranks = ranks[::-1]
    for i in range(8):
        file_text = COORDINATE_FONT.render(files[i], True, COORDINATE_COLOR); win.blit(file_text, (BUTTONS_WIDTH + MARGIN + i * SQUARE_SIZE + (SQUARE_SIZE - file_text.get_width()) / 2, HEIGHT - MARGIN - 60 + 5))
        rank_text = COORDINATE_FONT.render(ranks[i], True, COORDINATE_COLOR); win.blit(rank_text, (BUTTONS_WIDTH + MARGIN / 2 - rank_text.get_width() / 2, MARGIN + i * SQUARE_SIZE + (SQUARE_SIZE - rank_text.get_height()) / 2))

def draw_board(win, is_recording=False):
    """Desenează tabla cu culori diferite pentru recording vs live"""
    if is_recording:
        # Culori maro pentru recording
        dark_color, light_color = BROWN, LIGHT_BROWN
    else:
        # Culori verzi pentru joc live
        dark_color, light_color = LIVE_DARK, LIVE_LIGHT
    
    for r in range(8):
        for c in range(8):
            color = light_color if (r + c) % 2 == 0 else dark_color
            pygame.draw.rect(win, color, (BUTTONS_WIDTH + MARGIN + c * SQUARE_SIZE, MARGIN + r * SQUARE_SIZE, SQUARE_SIZE, SQUARE_SIZE))

def draw_pieces(win, board, images, selected_piece, flipped):
    for r_logic in range(8):
        for c_logic in range(8):
            row_on_screen = 7 - r_logic if flipped else r_logic; col_on_screen = 7 - c_logic if flipped else c_logic
            sq_idx = chess.square(c_logic, 7 - r_logic); piece = board.piece_at(sq_idx)
            if piece and sq_idx != selected_piece:
                p_sym, img_key = piece.symbol(), ('w' if piece.color == chess.WHITE else 'b') + piece.symbol().upper(); win.blit(images[img_key], (BUTTONS_WIDTH + MARGIN + col_on_screen * SQUARE_SIZE, MARGIN + row_on_screen * SQUARE_SIZE))

def get_square_from_mouse(pos, flipped):
    x, y = pos[0] - BUTTONS_WIDTH - MARGIN, pos[1] - MARGIN
    if not (0 <= x < BOARD_SIZE and 0 <= y < BOARD_SIZE): return None
    col_on_screen, row_on_screen = x // SQUARE_SIZE, y // SQUARE_SIZE
    col_logic, row_logic = (7 - col_on_screen, row_on_screen) if flipped else (col_on_screen, 7 - row_on_screen)
    if not (0 <= col_logic < 8 and 0 <= row_logic < 8): return None
    return chess.square(col_logic, row_logic)

def draw_multiline_text(surface, text, pos, font, color, max_width):
    words = text.split(' '); space = font.size(' ')[0]; x, y = pos; lines = []
    current_line = ""
    for word in words:
        if font.size(current_line + word)[0] < max_width: current_line += word + " "
        else: lines.append(current_line); current_line = word + " "
    lines.append(current_line)
    for line in lines: text_surface = font.render(line, True, color); surface.blit(text_surface, (x, y)); y += font.get_linesize()

def draw_history_panel(win, history, player_color):
    panel_y = MARGIN + BOARD_SIZE + 10; panel_height = HEIGHT - panel_y - 10; panel_rect = pygame.Rect(BUTTONS_WIDTH + MARGIN, panel_y, BOARD_SIZE, panel_height)
    pygame.draw.rect(win, HISTORY_BACKGROUND, panel_rect, border_radius=5)
    display_string = "";
    for i, move in enumerate(history):
        is_our_move = (i % 2 == 0) if player_color == chess.WHITE else (i % 2 != 0)
        prefix = "Eu: " if is_our_move else "El: "
        display_string += f"{prefix}{move} | "
    draw_multiline_text(win, display_string, (panel_rect.x + 10, panel_rect.y + 10), HISTORY_FONT, HISTORY_TEXT_COLOR, BOARD_SIZE - 20)
    copy_button_rect = pygame.Rect(panel_rect.right - 100, panel_rect.bottom - 35, 90, 25)
    pygame.draw.rect(win, UNDO_BUTTON_COLOR, copy_button_rect, border_radius=5)
    copy_text = COORDINATE_FONT.render("Copiază", True, WHITE_PIECE_COLOR); win.blit(copy_text, (copy_button_rect.x + (copy_button_rect.width - copy_text.get_width()) / 2, copy_button_rect.y + 5))
    return copy_button_rect

def draw_left_buttons(win, is_recording, undo_buttons_dict, record_button_rect, restart_button_rect, import_button_rect, import_folder_button_rect, text_input_name, active_input_name):
    """Desenează butoanele în stânga tablei și input-ul pentru nume în modul înregistrare"""
    # Fundal pentru zona de butoane
    buttons_rect = pygame.Rect(10, 10, BUTTONS_WIDTH - 20, HEIGHT - 20)
    pygame.draw.rect(win, (40, 40, 40), buttons_rect, border_radius=10)
    
    # Titlu pentru zona de butoane
    title_text = ASSISTANT_FONT.render("Comenzi:", True, WHITE_PIECE_COLOR)
    win.blit(title_text, (20, 30))
    
    y_pos = 80
    button_width = BUTTONS_WIDTH - 40
    button_height = 35  # Mai mici pentru a încăpea toate
    spacing = 45
    
    # 4 Butoane de Undo într-un grid 2x2
    undo_button_width = (button_width - 10) // 2  # Împarte pe 2 cu spațiu între ele
    undo_button_height = 30
    
    # Prima linie de butoane undo
    # Buton |< (la început)
    undo_buttons_dict['to_start'].update(20, y_pos, undo_button_width, undo_button_height)
    pygame.draw.rect(win, UNDO_BUTTON_COLOR, undo_buttons_dict['to_start'], border_radius=5)
    start_text = COORDINATE_FONT.render("|<", True, WHITE_PIECE_COLOR)
    text_x = undo_buttons_dict['to_start'].x + (undo_button_width - start_text.get_width()) // 2
    win.blit(start_text, (text_x, undo_buttons_dict['to_start'].y + 7))
    
    # Buton < (o mutare înapoi)
    undo_buttons_dict['one_back'].update(20 + undo_button_width + 10, y_pos, undo_button_width, undo_button_height)
    pygame.draw.rect(win, UNDO_BUTTON_COLOR, undo_buttons_dict['one_back'], border_radius=5)
    back_text = COORDINATE_FONT.render("<", True, WHITE_PIECE_COLOR)
    text_x = undo_buttons_dict['one_back'].x + (undo_button_width - back_text.get_width()) // 2
    win.blit(back_text, (text_x, undo_buttons_dict['one_back'].y + 7))
    
    # A doua linie de butoane undo
    y_pos_second = y_pos + undo_button_height + 5
    
    # Buton > (o mutare înainte)
    undo_buttons_dict['one_forward'].update(20, y_pos_second, undo_button_width, undo_button_height)
    # Culoare diferită dacă nu putem merge înainte
    forward_color = (60, 60, 120) if 'move_history_forward' not in globals() or not move_history_forward else UNDO_BUTTON_COLOR
    pygame.draw.rect(win, forward_color, undo_buttons_dict['one_forward'], border_radius=5)
    forward_text = COORDINATE_FONT.render(">", True, WHITE_PIECE_COLOR)
    text_x = undo_buttons_dict['one_forward'].x + (undo_button_width - forward_text.get_width()) // 2
    win.blit(forward_text, (text_x, undo_buttons_dict['one_forward'].y + 7))
    
    # Buton >| (la sfârșit)
    undo_buttons_dict['to_end'].update(20 + undo_button_width + 10, y_pos_second, undo_button_width, undo_button_height)
    end_color = (60, 60, 120) if 'move_history_forward' not in globals() or not move_history_forward else UNDO_BUTTON_COLOR
    pygame.draw.rect(win, end_color, undo_buttons_dict['to_end'], border_radius=5)
    end_text = COORDINATE_FONT.render(">|", True, WHITE_PIECE_COLOR)
    text_x = undo_buttons_dict['to_end'].x + (undo_button_width - end_text.get_width()) // 2
    win.blit(end_text, (text_x, undo_buttons_dict['to_end'].y + 7))
    
    # Actualizează y_pos pentru butoanele următoare
    y_pos = y_pos_second + undo_button_height + 15
    
    # Butonul Record
    record_button_text = "Salvează Linia" if is_recording else "Înregistrează Linie"
    record_button_color = RECORDING_BUTTON_COLOR if is_recording else RECORD_BUTTON_COLOR
    record_button_rect.update(20, y_pos, button_width, button_height)
    pygame.draw.rect(win, record_button_color, record_button_rect, border_radius=8)
    record_text = COORDINATE_FONT.render(record_button_text, True, WHITE_PIECE_COLOR)
    text_x = record_button_rect.x + (button_width - record_text.get_width()) // 2
    win.blit(record_text, (text_x, record_button_rect.y + 10))
    y_pos += spacing
    
    # Butonul Import SINGULAR
    import_button_rect.update(20, y_pos, button_width, 30)  # Mai mic
    pygame.draw.rect(win, (100, 150, 100), import_button_rect, border_radius=8)
    import_text = COORDINATE_FONT.render("Import PGN", True, WHITE_PIECE_COLOR)
    text_x = import_button_rect.x + (button_width - import_text.get_width()) // 2
    win.blit(import_text, (text_x, import_button_rect.y + 7))
    y_pos += 35
    
    # BUTONUL NOU - Import FOLDER
    import_folder_button_rect.update(20, y_pos, button_width, 30)
    pygame.draw.rect(win, (150, 100, 150), import_folder_button_rect, border_radius=8)  # Culoare violet
    folder_text = COORDINATE_FONT.render("Import Folder PGN", True, WHITE_PIECE_COLOR)
    text_x = import_folder_button_rect.x + (button_width - folder_text.get_width()) // 2
    win.blit(folder_text, (text_x, import_folder_button_rect.y + 7))
    y_pos += 40
    
    # Butonul Restart
    restart_button_rect.update(20, y_pos, button_width, button_height)
    pygame.draw.rect(win, RESTART_BUTTON_COLOR, restart_button_rect, border_radius=8)
    restart_text = COORDINATE_FONT.render("Meniu Principal", True, WHITE_PIECE_COLOR)
    text_x = restart_button_rect.x + (button_width - restart_text.get_width()) // 2
    win.blit(restart_text, (text_x, restart_button_rect.y + 10))
    
    # Informații de înregistrare și input pentru nume dacă e cazul
    if is_recording:
        info_y = y_pos + spacing + 20
        rec_info = COORDINATE_FONT.render("MOD ÎNREGISTRARE ACTIV", True, RECORDING_BUTTON_COLOR)
        win.blit(rec_info, (20, info_y))
        
        # Input pentru numele capcanei
        name_y = info_y + 40
        name_label = COORDINATE_FONT.render("Nume Capcană:", True, WHITE_PIECE_COLOR)
        win.blit(name_label, (20, name_y))
        
        # Input box pentru nume
        input_y = name_y + 25
        text_input_name.font_color = WHITE_PIECE_COLOR if active_input_name else (150, 150, 150)
        win.blit(text_input_name.surface, (20, input_y))


def draw_message_area(win, current_turn, player_color, is_recording, board):
    """Desenează zona de mesaje în partea de jos"""
    message_y = HEIGHT - 50
    message_width = BUTTONS_WIDTH + BOARD_SIZE + MARGIN
    message_rect = pygame.Rect(BUTTONS_WIDTH, message_y, message_width, 40)
    
    # Fundal pentru zona de mesaje
    pygame.draw.rect(win, MESSAGE_BACKGROUND, message_rect, border_radius=5)
    pygame.draw.rect(win, (100, 100, 100), message_rect, 1, border_radius=5)
    
    # Determină mesajul
    if is_recording:
        message = "MOD ÎNREGISTRARE - Joacă ambele culori pentru a crea capcana"
        color = RECORDING_BUTTON_COLOR
    elif board.is_checkmate():
        winner = "Albul" if current_turn == chess.BLACK else "Negrul"
        message = f"Mat! {winner} a câștigat!"
        color = (255, 100, 100)
    elif board.is_stalemate():
        message = "Remiză prin pat!"
        color = (255, 255, 100)
    elif board.is_check():
        if current_turn == player_color:
            message = "Ești în șah! Trebuie să muți."
            color = (255, 150, 150)
        else:
            message = "Adversarul este în șah. Așteaptă mutarea lui."
            color = (150, 255, 150)
    elif current_turn == player_color:
        message = "Este rândul tău să muți"
        color = (150, 255, 150)
    else:
        message = "Rândul adversarului - așteaptă mutarea"
        color = (255, 255, 150)
    
    # Desenează textul
    message_text = COORDINATE_FONT.render(message, True, color)
    text_x = message_rect.x + (message_rect.width - message_text.get_width()) // 2
    text_y = message_rect.y + (message_rect.height - message_text.get_height()) // 2
    win.blit(message_text, (text_x, text_y))

def draw_database_stats(win, total_traps, player_color):
    """Desenează statisticile bazei de date în partea de sus a sidebar-ului"""
    sidebar_x = BUTTONS_WIDTH + BOARD_SIZE + 2 * MARGIN + 10
    sidebar_y = 20
    sidebar_width = SIDEBAR_WIDTH - 30
    
    # Fundal pentru statistici
    stats_rect = pygame.Rect(sidebar_x, sidebar_y, sidebar_width, 80)
    pygame.draw.rect(win, (30, 30, 40), stats_rect, border_radius=8)
    pygame.draw.rect(win, (80, 80, 80), stats_rect, 2, border_radius=8)
    
    # Titlu
    color_name = "Albele" if player_color == chess.WHITE else "Negrele"
    title_text = COORDINATE_FONT.render(f"Baza de Date - {color_name}", True, WHITE_PIECE_COLOR)
    win.blit(title_text, (sidebar_x + 10, sidebar_y + 10))
    
    # Numărul de capcane
    count_text = ASSISTANT_FONT.render(f"{total_traps} capcane", True, ASSISTANT_TEXT_COLOR)
    win.blit(count_text, (sidebar_x + 10, sidebar_y + 35))
    
    # Informații suplimentare
    info_text = COORDINATE_FONT.render(f"Din fișierul traps.xlsx", True, (180, 180, 180))
    win.blit(info_text, (sidebar_x + 10, sidebar_y + 60))

def draw_traps_sidebar(win, suggestions, highlighted_move, opponent_target_info, flipped, expected_opponent_move=None, scroll_offset=0, is_recording=False, total_traps=0, player_color=chess.WHITE):
    """Desenează sidebar-ul cu capcane și scroll"""
    # Desenează statisticile bazei de date
    draw_database_stats(win, total_traps, player_color)
    
    draw_database_stats(win, total_traps, player_color)
    
    sidebar_x = BUTTONS_WIDTH + BOARD_SIZE + 2 * MARGIN + 10
    sidebar_y = 110  # MODIFICAT: Mutăm mai jos pentru statistici
    sidebar_width = SIDEBAR_WIDTH - 30
    sidebar_height = HEIGHT - 240  # MODIFICAT: Reducem înălțimea
        
    # Fundal pentru sidebar
    sidebar_rect = pygame.Rect(sidebar_x, sidebar_y, sidebar_width, sidebar_height)
    pygame.draw.rect(win, (45, 45, 45), sidebar_rect, border_radius=10)
    pygame.draw.rect(win, (100, 100, 100), sidebar_rect, 2, border_radius=10)
    
    # Nu afișa nimic în modul înregistrare
    if is_recording:
        return [], scroll_offset
    
    # Titlul pentru lista de capcane
    title_text = ASSISTANT_FONT.render("Capcane Disponibile:", True, WHITE_PIECE_COLOR)
    win.blit(title_text, (sidebar_x + 10, sidebar_y + 10))
    
    # Zona pentru lista de capcane
    list_y = sidebar_y + 40  # MODIFICAT: Ajustăm poziția
    list_height = sidebar_height - 50  # MODIFICAT: Ajustăm înălțimea
    
    suggestion_rects = []
    
    if suggestions:
        item_height = 35
        visible_items = list_height // item_height
        total_items = len(suggestions)
        
        # Calculează scroll-ul maxim
        max_scroll = max(0, total_items - visible_items)
        scroll_offset = max(0, min(scroll_offset, max_scroll))
        
        # Desenează bara de scroll dacă e nevoie
        if total_items > visible_items:
            scrollbar_x = sidebar_x + sidebar_width - 20
            scrollbar_height = list_height
            scrollbar_thumb_height = max(20, (visible_items / total_items) * scrollbar_height)
            scrollbar_thumb_y = list_y + (scroll_offset / max_scroll) * (scrollbar_height - scrollbar_thumb_height)
            
            # Fundalul barei de scroll
            pygame.draw.rect(win, (60, 60, 60), (scrollbar_x, list_y, 15, scrollbar_height))
            # Thumb-ul barei de scroll
            pygame.draw.rect(win, (120, 120, 120), (scrollbar_x, scrollbar_thumb_y, 15, scrollbar_thumb_height), border_radius=7)
        
        # Desenează elementele vizibile
        for i in range(scroll_offset, min(scroll_offset + visible_items, total_items)):
            y_pos = list_y + (i - scroll_offset) * item_height
            trap = suggestions[i]
            
            # Fundalul pentru fiecare element
            item_rect = pygame.Rect(sidebar_x + 10, y_pos, sidebar_width - 40, item_height - 2)
            pygame.draw.rect(win, (60, 60, 60), item_rect, border_radius=5)
            
            # Textul capcanei
            trap_text_str = f"{i+1}. {trap['name']} ({trap['suggestion']})"
            if trap.get('opponent_target_san'):
                trap_text_str += f" → {trap['opponent_target_san']}"
            
            # Taie textul dacă e prea lung
            if len(trap_text_str) > 45:
                trap_text_str = trap_text_str[:42] + "..."
                
            trap_text = COORDINATE_FONT.render(trap_text_str, True, ASSISTANT_TEXT_COLOR)
            win.blit(trap_text, (sidebar_x + 15, y_pos + 8))
            
            suggestion_rects.append(item_rect)
    
    elif opponent_target_info:
        wait_text = ASSISTANT_FONT.render("Aștept răspuns...", True, WHITE_PIECE_COLOR)
        win.blit(wait_text, (sidebar_x + 10, list_y))
        
        target_san = opponent_target_info.get('opponent_target_san', 'N/A')
        target_text = COORDINATE_FONT.render(f"Ținta: {target_san}", True, ASSISTANT_TEXT_COLOR)
        win.blit(target_text, (sidebar_x + 10, list_y + 40))
    
    else:
        no_trap_text = ASSISTANT_FONT.render("Ești pe cont propriu.", True, WHITE_PIECE_COLOR)
        win.blit(no_trap_text, (sidebar_x + 10, list_y))
    
    # Evidențierile pe tablă (mutate aici pentru organizare)
    if highlighted_move:
        highlight_surface = pygame.Surface((SQUARE_SIZE, SQUARE_SIZE), pygame.SRCALPHA)
        highlight_surface.fill(HIGHLIGHT_COLOR)
        from_file, from_rank = chess.square_file(highlighted_move.from_square), chess.square_rank(highlighted_move.from_square)
        from_col, from_row = (7 - from_file, from_rank) if flipped else (from_file, 7 - from_rank)
        win.blit(highlight_surface, (BUTTONS_WIDTH + MARGIN + from_col * SQUARE_SIZE, MARGIN + from_row * SQUARE_SIZE))
        to_file, to_rank = chess.square_file(highlighted_move.to_square), chess.square_rank(highlighted_move.to_square)
        to_col, to_row = (7 - to_file, to_rank) if flipped else (to_file, 7 - to_rank)
        win.blit(highlight_surface, (BUTTONS_WIDTH + MARGIN + to_col * SQUARE_SIZE, MARGIN + to_row * SQUARE_SIZE))
    
    if 'all_expected_moves' in globals() and all_expected_moves:
        expected_surface = pygame.Surface((SQUARE_SIZE, SQUARE_SIZE), pygame.SRCALPHA)
        expected_surface.fill(EXPECTED_MOVE_COLOR)
        
        for move in all_expected_moves:
            # Evidențiază pătratul de plecare
            from_file, from_rank = chess.square_file(move.from_square), chess.square_rank(move.from_square)
            from_col, from_row = (7 - from_file, from_rank) if flipped else (from_file, 7 - from_rank)
            win.blit(expected_surface, (BUTTONS_WIDTH + MARGIN + from_col * SQUARE_SIZE, MARGIN + from_row * SQUARE_SIZE))
            
            # Evidențiază pătratul de sosire
            to_file, to_rank = chess.square_file(move.to_square), chess.square_rank(move.to_square)
            to_col, to_row = (7 - to_file, to_rank) if flipped else (to_file, 7 - to_rank)
            win.blit(expected_surface, (BUTTONS_WIDTH + MARGIN + to_col * SQUARE_SIZE, MARGIN + to_row * SQUARE_SIZE))
    
    elif expected_opponent_move:
        # Fallback pentru o singură mutare
        expected_surface = pygame.Surface((SQUARE_SIZE, SQUARE_SIZE), pygame.SRCALPHA)
        expected_surface.fill(EXPECTED_MOVE_COLOR)
        from_file, from_rank = chess.square_file(expected_opponent_move.from_square), chess.square_rank(expected_opponent_move.from_square)
        from_col, from_row = (7 - from_file, from_rank) if flipped else (from_file, 7 - from_rank)
        win.blit(expected_surface, (BUTTONS_WIDTH + MARGIN + from_col * SQUARE_SIZE, MARGIN + from_row * SQUARE_SIZE))
        to_file, to_rank = chess.square_file(expected_opponent_move.to_square), chess.square_rank(expected_opponent_move.to_square)
        to_col, to_row = (7 - to_file, to_rank) if flipped else (to_file, 7 - to_rank)
        win.blit(expected_surface, (BUTTONS_WIDTH + MARGIN + to_col * SQUARE_SIZE, MARGIN + to_row * SQUARE_SIZE))    
    return suggestion_rects, scroll_offset

def import_folder_traps_from_pgn(folder_path, max_moves=25):
    """Importă capcane din toate fișierele PGN dintr-un folder"""
    import chess.pgn
    import os
    
    white_traps_all = []
    black_traps_all = []
    processed_files = 0
    failed_files = []
    
    # Găsește toate fișierele PGN din folder
    pgn_files = []
    for file in os.listdir(folder_path):
        if file.lower().endswith('.pgn'):
            pgn_files.append(os.path.join(folder_path, file))
    
    if not pgn_files:
        print(f"[IMPORT MASĂ] Nu am găsit fișiere PGN în {folder_path}")
        return [], []
    
    print(f"[IMPORT MASĂ] Găsite {len(pgn_files)} fișiere PGN pentru procesare...")
    
    for pgn_file_path in pgn_files:
        try:
            print(f"[IMPORT MASĂ] Procesez: {os.path.basename(pgn_file_path)}")
            
            with open(pgn_file_path, 'r', encoding='utf-8') as pgn_file:
                game_count_file = 0
                processed_count_file = 0
                
                while True:
                    game = chess.pgn.read_game(pgn_file)
                    if game is None:
                        break
                    
                    game_count_file += 1
                    moves = list(game.mainline_moves())
                    
                    # Filtrează după lungime
                    if 4 <= len(moves) <= max_moves:
                        # Convertește mutările în notație SAN
                        board = chess.Board()
                        san_moves = []
                        
                        try:
                            for move in moves:
                                san_moves.append(board.san(move))
                                board.push(move)
                            
                            # Creează numele capcanei cu numele fișierului
                            file_name = os.path.splitext(os.path.basename(pgn_file_path))[0]
                            trap_name = game.headers.get("ChapterName", 
                                       game.headers.get("Event", 
                                       f"{file_name} - Trapă {game_count_file}"))
                            
                            # Verifică dacă se termină cu mat
                            is_checkmate = "#" in san_moves[-1] if san_moves else False
                            if is_checkmate:
                                trap_name += " (Mat)"
                            
                            # Determină cine câștigă
                            total_moves = len(san_moves)
                            if total_moves > 0:
                                if is_checkmate:
                                    if total_moves % 2 == 1:  # Albele au mutat ultimul
                                        white_traps_all.append({
                                            "name": trap_name,
                                            "moves": san_moves,
                                            "result": "1-0"
                                        })
                                        processed_count_file += 1
                                    else:  # Negrele au mutat ultimul
                                        black_traps_all.append({
                                            "name": trap_name,
                                            "moves": san_moves,
                                            "result": "0-1"
                                        })
                                        processed_count_file += 1
                                else:
                                    # Pentru trap-uri non-mat, adaugă la albe (majoritatea sunt pentru albe)
                                    white_traps_all.append({
                                        "name": trap_name,
                                        "moves": san_moves,
                                        "result": "1-0"
                                    })
                                    processed_count_file += 1
                        
                        except Exception as e:
                            continue  # Ignore invalid games
                
                print(f"[IMPORT MASĂ] {os.path.basename(pgn_file_path)}: {processed_count_file}/{game_count_file} partide procesate")
                processed_files += 1
                
        except Exception as e:
            print(f"[IMPORT MASĂ] EROARE la {os.path.basename(pgn_file_path)}: {e}")
            failed_files.append(os.path.basename(pgn_file_path))
            continue
    
    print(f"\n[IMPORT MASĂ] SUMAR:")
    print(f"- Fișiere procesate cu succes: {processed_files}/{len(pgn_files)}")
    print(f"- Capcane găsite pentru albe: {len(white_traps_all)}")
    print(f"- Capcane găsite pentru negre: {len(black_traps_all)}")
    
    if failed_files:
        print(f"- Fișiere cu erori: {', '.join(failed_files)}")
    
    return white_traps_all, black_traps_all    


# --- PUNCTUL DE INTRARE ÎN PROGRAM ---

# Încarcă setările la pornire
settings = load_settings()
print(f"[SETĂRI] Limite PGN: {settings['pgn_import_max_moves']} semi-mutări")

player_choice = start_screen()
trap_library = {}
total_traps_count = 0  # ADAUGĂ această variabilă globală
if player_choice is not None:
    trap_library, total_traps_count = build_trap_library_from_excel(player_choice)  # MODIFICĂ apelul
    is_flipped = (player_choice == chess.BLACK)
    board = chess.Board()
    piece_images = load_piece_images()
    
    # Inițializare variabile pentru butoane (folosind pygame.Rect pentru poziționare dinamică)
    undo_buttons_dict = {
    'to_start': pygame.Rect(0, 0, 0, 0),    # |
    'one_back': pygame.Rect(0, 0, 0, 0),    # 
    'one_forward': pygame.Rect(0, 0, 0, 0), # >
    'to_end': pygame.Rect(0, 0, 0, 0)       # >|
    }

    # Pentru a permite navigarea înainte
    move_history_forward = []  # Stack pentru mutările "refăcute"

    record_button_rect = pygame.Rect(0, 0, 0, 0)
    restart_button_rect = pygame.Rect(0, 0, 0, 0)
    import_button_rect = pygame.Rect(0, 0, 0, 0)
    import_folder_button_rect = pygame.Rect(0, 0, 0, 0)  # ACEASTĂ LINIE E NOUĂ
    copy_history_button_rect = None
    
    selected_piece, selected_piece_image = None, None
    trap_suggestions, trap_suggestions_rects, highlighted_move, last_fen_checked = [], [], None, None
    opponent_target_info, last_player_move_san = None, None
    expected_opponent_move = None
    is_recording, move_history_for_recording = False, []
    scroll_offset = 0  # Pentru scroll-ul în lista de capcane
    
    manager_name = pygame_textinput.TextInputManager(validator = lambda input: len(input) <= 30)
    text_input_name = pygame_textinput.TextInputVisualizer(manager=manager_name, font_object=COORDINATE_FONT, font_color=WHITE_PIECE_COLOR, cursor_color=WHITE_PIECE_COLOR)
    active_input_name = True  # Nouă variabilă pentru input-ul nume

    running = True; clock = pygame.time.Clock()
    while running:
        events = pygame.event.get()
        if is_recording: text_input_name.update(events)
        
        clock.tick(60); mouse_pos = pygame.mouse.get_pos()
        current_fen = board.fen()
        if current_fen != last_fen_checked:
            print(f"\n[SISTEM] Tabla s-a schimbat. Verific FEN: {current_fen}")
            trap_suggestions, highlighted_move, opponent_target_info = [], None, None
            expected_opponent_move = None
            all_expected_moves = []  # NOUĂ: listă pentru toate mutările așteptate
            
            if board.turn == player_choice and not is_recording:
                library_entry = trap_library.get(current_fen)
                if library_entry and "player_options" in library_entry:
                    # Obține mutările curente și sortează capcanele
                    current_moves = get_current_position_moves(board)
                    unsorted_suggestions = library_entry["player_options"]
                    trap_suggestions = sort_traps_by_common_beginning(unsorted_suggestions, current_moves)
                    print(f"[ASISTENT] Ai {len(trap_suggestions)} opțiune(i) de capcană (sortate după relevanță).")
                else: 
                    print("[ASISTENT] Ești pe cont propriu.")
                
            elif not is_recording and last_player_move_san and board.move_stack:
                # NOUĂ LOGICĂ: Verifică dacă să facă pauză după prima mutare
                move_count = len(board.move_stack)
                
                # Dacă joci cu albele și tocmai ai făcut prima mutare (move_count = 1)
                # sau dacă joci cu negrele și tocmai ai făcut prima mutare (move_count = 2)
                should_pause = False
                if player_choice == chess.WHITE and move_count == 1:
                    should_pause = True
                    print("[PAUZĂ] Prima mutare albă completă - pauză în sugerări pentru următoarea mutare a negrului.")
                elif player_choice == chess.BLACK and move_count == 2:
                    should_pause = True
                    print("[PAUZĂ] Prima mutare neagră completă - pauză în sugerări pentru următoarea mutare a albului.")
                
                if not should_pause:
                    temp_board = board.copy(); temp_board.pop(); fen_before_our_move = temp_board.fen()
                    library_entry = trap_library.get(fen_before_our_move)
                    if library_entry and "player_options" in library_entry:
                        matching_traps = []  # Colectează toate capcanele care se potrivesc
                        for option in library_entry['player_options']:
                            if last_player_move_san == option['suggestion']:
                                matching_traps.append(option)
                                # Calculează mutarea așteptată pentru evidențiere
                                if option.get('opponent_target_san'):
                                    try:
                                        move = board.parse_san(option['opponent_target_san'])
                                        all_expected_moves.append(move)
                                        print(f"[ASISTENT] Capcană activă: {option.get('trap_name', 'Fără nume')} - Ținta: {option['opponent_target_san']}")
                                    except ValueError:
                                        print(f"[EROARE] Nu pot parsa mutarea: {option['opponent_target_san']}")
                        
                        if matching_traps:
                            # Folosește prima capcană pentru informații, dar evidențiază toate mutările
                            opponent_target_info = matching_traps[0]
                            print(f"[ASISTENT] {len(matching_traps)} capcană/e activă/e! Evidențiez toate țintele.")
                else:
                    print("[PAUZĂ] Sugerările sunt temporar dezactivate - așteaptă următoarea mutare.")
            last_fen_checked = current_fen

        for event in events:
            if event.type == pygame.QUIT: running = False
            elif event.type == pygame.MOUSEWHEEL:
                # Gestionează scroll-ul în lista de capcane
                if trap_suggestions:  # Doar dacă avem capcane de afișat
                    scroll_offset += -event.y  # Inversează direcția pentru scroll natural
                    scroll_offset = max(0, scroll_offset)  # Nu permite scroll negativ
            elif event.type == pygame.MOUSEBUTTONDOWN:
                # Gestionarea butoanelor de navigare
                if undo_buttons_dict['to_start'].collidepoint(event.pos):
                    # Mergi la începutul partidei
                    while board.move_stack:
                        if move_history_for_recording: move_history_for_recording.pop()
                        move_history_forward.append(board.pop())
                    last_fen_checked = None
                    print("\n[NAV] Revenit la început.")
                    continue
                elif undo_buttons_dict['one_back'].collidepoint(event.pos):
                    # O mutare înapoi
                    if board.move_stack:
                        if move_history_for_recording: move_history_for_recording.pop()
                        move_history_forward.append(board.pop())
                        last_fen_checked = None
                        print("\n[NAV] O mutare înapoi.")
                    else: 
                        print("[NAV] Nu mai sunt mutări de anulat.")
                    continue
                elif undo_buttons_dict['one_forward'].collidepoint(event.pos):
                    # O mutare înainte
                    if move_history_forward:
                        move = move_history_forward.pop()
                        board.push(move)
                        if is_recording: 
                            move_history_for_recording.append(board.san(move))
                        last_fen_checked = None
                        print("\n[NAV] O mutare înainte.")
                    else: 
                        print("[NAV] Nu mai sunt mutări de refăcut.")
                    continue
                elif undo_buttons_dict['to_end'].collidepoint(event.pos):
                    # Mergi la sfârșitul partidei
                    while move_history_forward:
                        move = move_history_forward.pop()
                        board.push(move)
                        if is_recording: 
                            move_history_for_recording.append(board.san(move))
                    last_fen_checked = None
                    print("\n[NAV] Ajuns la sfârșit.")
                    continue
                if copy_history_button_rect and copy_history_button_rect.collidepoint(event.pos):
                    excel_string = "\t".join(move_history_for_recording); pyperclip.copy(excel_string); print(f"\n[CLIPBOARD] Linia a fost copiată: {excel_string}"); continue
                if record_button_rect and record_button_rect.collidepoint(event.pos):
                    is_recording = not is_recording
                    if is_recording:
                        move_history_for_recording = []; board.reset(); last_fen_checked = None; trap_suggestions = []; opponent_target_info = None
                        move_history_forward = []  # ACEASTĂ LINIE E NOUĂ
                        print("\n[REC] Mod Înregistrare ACTIVAT.")
                    else:
                        trap_name = text_input_name.value if text_input_name.value else "Linie Fără Nume"
                        if move_history_for_recording:
                            sheet_name = "White" if player_choice == chess.WHITE else "Black"
                            row_idx = get_next_free_row(sheet_name)  # Auto-detectează rândul
                            if write_new_line_to_excel(sheet_name, row_idx, trap_name, move_history_for_recording): 
                                trap_library, total_traps_count = build_trap_library_from_excel(player_choice)
                        else: 
                            print("[EROARE SALVARE] Nicio mutare de salvat.")
                        text_input_name.value = ""; board.reset(); last_fen_checked = None
                        move_history_forward = []  # ACEASTĂ LINIE E NOUĂ
                        print("[REC] Mod Înregistrare OPRIT.")
                    continue
                if restart_button_rect and restart_button_rect.collidepoint(event.pos):
                    print("\n[RESTART] Revin la meniul principal...")
                    running = False
                    player_choice = start_screen()
                    if player_choice is not None:
                        trap_library, total_traps_count = build_trap_library_from_excel(player_choice)
                        is_flipped = (player_choice == chess.BLACK)
                        board = chess.Board()
                        selected_piece, selected_piece_image = None, None
                        trap_suggestions, trap_suggestions_rects, highlighted_move, last_fen_checked = [], [], None, None
                        opponent_target_info, last_player_move_san = None, None
                        expected_opponent_move = None
                        is_recording, move_history_for_recording = False, []
                        scroll_offset = 0
                        move_history_forward = []  # ACEASTĂ LINIE E NOUĂ
                        text_input_name.value = ""
                        active_input_name = True
                        running = True
                    continue

                if import_button_rect and import_button_rect.collidepoint(event.pos):
                    print("\n[IMPORT] Selectează fișierul PGN...")
                    import tkinter.filedialog as fd
                    import tkinter.simpledialog as sd
                    import os
                    
                    # Încarcă setările
                    settings = load_settings()
                    last_directory = settings.get("last_pgn_directory", "")
                    current_max_moves = settings.get("pgn_import_max_moves", 25)
                    
                    pgn_file = fd.askopenfilename(
                        title="Selectează fișierul PGN",
                        initialdir=last_directory if last_directory else None,
                        filetypes=[("PGN files", "*.pgn"), ("All files", "*.*")]
                    )
                    
                    if pgn_file:
                        # Salvează directorul pentru următoarea dată
                        update_setting("last_pgn_directory", os.path.dirname(pgn_file))
                        
                        # Întreabă utilizatorul pentru limita de semi-mutări
                        max_moves = sd.askinteger(
                            "Limită Semi-mutări", 
                            f"Câte semi-mutări maxim să accepte?\n"
                            f"(o semi-mutare = o mișcare de alb SAU negru)\n"
                            f"Exemplu: 25 semi-mutări = ~12 mutări complete\n"
                            f"Ultima setare: {current_max_moves}",
                            initialvalue=current_max_moves,
                            minvalue=5,
                            maxvalue=100
                        )
                        
                        if max_moves:
                            # Salvează noua setare dacă s-a schimbat
                            if max_moves != current_max_moves:
                                update_setting("pgn_import_max_moves", max_moves)
                                print(f"[SETĂRI] Limita de semi-mutări salvată: {max_moves}")
                            
                            white_traps, black_traps = import_traps_from_pgn(pgn_file, max_moves)
                            if white_traps or black_traps:
                                if save_traps_to_excel(white_traps, black_traps):
                                    # Reîncarcă biblioteca de capcane
                                    trap_library, total_traps_count = build_trap_library_from_excel(player_choice)
                                    print("[IMPORT] Biblioteca de capcane a fost actualizată!")
                    continue


                if import_folder_button_rect and import_folder_button_rect.collidepoint(event.pos):
                    print("\n[IMPORT MASĂ] Selectează folderul cu PGN-uri...")
                    import tkinter.filedialog as fd
                    import tkinter.simpledialog as sd
                    import os
                    
                    # Încarcă setările
                    settings = load_settings()
                    last_directory = settings.get("last_pgn_directory", "")
                    current_max_moves = settings.get("pgn_import_max_moves", 25)
                    
                    folder_path = fd.askdirectory(
                        title="Selectează folderul cu fișiere PGN",
                        initialdir=last_directory if last_directory else None
                    )
                    
                    if folder_path:
                        # Salvează directorul pentru următoarea dată
                        update_setting("last_pgn_directory", folder_path)
                        
                        # Întreabă utilizatorul pentru limita de semi-mutări
                        max_moves = sd.askinteger(
                            "Limită Semi-mutări (Import Masă)", 
                            f"Câte semi-mutări maxim să accepte pentru TOATE fișierele?\n"
                            f"(o semi-mutare = o mișcare de alb SAU negru)\n"
                            f"Exemplu: 25 semi-mutări = ~12 mutări complete\n"
                            f"Ultima setare: {current_max_moves}",
                            initialvalue=current_max_moves,
                            minvalue=5,
                            maxvalue=100
                        )
                        
                        if max_moves:
                            # Salvează noua setare dacă s-a schimbat
                            if max_moves != current_max_moves:
                                update_setting("pgn_import_max_moves", max_moves)
                                print(f"[SETĂRI] Limita de semi-mutări salvată: {max_moves}")
                            
                            # IMPORT ÎN MASĂ
                            white_traps, black_traps = import_folder_traps_from_pgn(folder_path, max_moves)
                            if white_traps or black_traps:
                                if save_traps_to_excel(white_traps, black_traps):
                                    # Reîncarcă biblioteca de capcane
                                    trap_library, total_traps_count = build_trap_library_from_excel(player_choice)
                                    print("[IMPORT MASĂ] Biblioteca de capcane a fost actualizată!")
                    continue

            
                clicked_on_suggestion = False
                if not is_recording:
                    for i, rect in enumerate(trap_suggestions_rects):
                        if rect.collidepoint(event.pos):
                            suggestion_str = trap_suggestions[i]['suggestion']
                            try: highlighted_move = board.parse_san(suggestion_str)
                            except Exception: highlighted_move = None
                            clicked_on_suggestion = True; break
                if not clicked_on_suggestion:
                    start_square = get_square_from_mouse(mouse_pos, is_flipped)
                    if start_square is not None:
                        piece = board.piece_at(start_square)
                        if piece and piece.color == board.turn:
                            selected_piece, highlighted_move = start_square, None
                            image_key = ('w' if piece.color == chess.WHITE else 'b') + piece.symbol().upper(); selected_piece_image = piece_images[image_key]
            elif event.type == pygame.MOUSEBUTTONUP:
                if selected_piece is not None:
                    end_square = get_square_from_mouse(mouse_pos, is_flipped)
                    if end_square is not None:
                        move = chess.Move(selected_piece, end_square)
                        if board.piece_type_at(selected_piece) == chess.PAWN and chess.square_rank(end_square) in [0, 7]: move.promotion = chess.QUEEN
                        if move in board.legal_moves:
                            original_move_san = board.san(move)
                            if board.turn == player_choice: last_player_move_san = original_move_san
                            else: last_player_move_san = None
                            board.push(move)
                            if is_recording: move_history_for_recording.append(original_move_san)
                            print(f"\n[MUTARE] {chess.COLOR_NAMES[not board.turn].upper()} a mutat: {board.peek().uci()}.")
                    selected_piece, selected_piece_image = None, None
        
        WIN.fill(BACKGROUND_COLOR)
        draw_board(WIN, is_recording); draw_coordinates(WIN, is_flipped)
        draw_pieces(WIN, board, piece_images, selected_piece, is_flipped)
        
        # Desenează butoanele în stânga
        draw_left_buttons(WIN, is_recording, undo_buttons_dict, record_button_rect, restart_button_rect, import_button_rect, import_folder_button_rect, text_input_name, active_input_name)
        
        # Desenează sidebar-ul cu capcane și scroll
        trap_suggestions_rects, scroll_offset = draw_traps_sidebar(WIN, trap_suggestions, highlighted_move, opponent_target_info, is_flipped, expected_opponent_move if 'all_expected_moves' not in locals() or not all_expected_moves else None, scroll_offset, is_recording, total_traps_count, player_choice)
        
        copy_history_button_rect = draw_history_panel(WIN, move_history_for_recording, player_choice)
        draw_message_area(WIN, board.turn, player_choice, is_recording, board)
        
        if selected_piece_image: WIN.blit(selected_piece_image, (mouse_pos[0] - SQUARE_SIZE / 2, mouse_pos[1] - SQUARE_SIZE / 2))
        pygame.display.flip()

pygame.quit()
print("Program încheiat.")